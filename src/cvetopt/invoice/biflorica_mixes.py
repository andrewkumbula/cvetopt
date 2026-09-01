"""
Разбор миксов: заполненный «Шаблон ДД.ММ.ГГ» + Biflorica Mix →
отдельные позиции с средневзвешенной ценой (только длины 50 и 60).
"""

from __future__ import annotations

import re
import shutil
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from cvetopt.invoice.biflorica_split import (
    biflorica_deals_header_or_raise,
    find_latest_biflorica_report,
    is_split_output_name,
)
from cvetopt.invoice.xlsx_read import (
    ensure_xlsx_workbook,
    grid_by_row,
    pick_data_worksheet,
    read_excel_grid,
)

LogFn = Callable[[str], None]

# По инструкции в шаблоне: отделяем только от Микс 50 и 60 см.
_TARGET_LENGTHS = ("50", "60")
_FILLED_TEMPLATE_STEM_RE = re.compile(
    r"^ша[бю]лон\s+.+$",  # шаблон / Шаблон / шаюлон (опечатки)
    re.IGNORECASE,
)


def _default_log(_msg: str) -> None:
    pass


def _norm(text: object) -> str:
    return " ".join(str(text or "").split()).strip()


def _as_float(value: object) -> float | None:
    text = _norm(value).replace(",", ".")
    if not text or text in {"-", "—"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_qty(value: object) -> int:
    num = _as_float(value)
    if num is None or num <= 0:
        return 0
    return int(round(num))


@dataclass(frozen=True)
class TemplateLine:
    excel_row: int
    code: str  # C — англ. код (Mix R, Mondial, …)
    title_ru: str  # D
    qtys: dict[str, int]  # длина → кол-во


@dataclass(frozen=True)
class TemplateDemand:
    path: Path
    lines: tuple[TemplateLine, ...]
    totals: dict[str, int]  # длина → итог


@dataclass
class MixTake:
    bif_row: int
    take_qty: int
    price: float
    plantation: str


@dataclass
class LengthPlan:
    length: str
    need: int
    lines: list[TemplateLine] = field(default_factory=list)
    takes: list[MixTake] = field(default_factory=list)
    avg_price: float = 0.0


def is_filled_sklad_template_name(name: str) -> bool:
    stem = Path(name).stem
    if stem.casefold() == "шаблон":
        return False
    return bool(_FILLED_TEMPLATE_STEM_RE.match(stem))


def find_latest_filled_sklad_template(sklad_dir: Path) -> Path:
    if not sklad_dir.is_dir():
        raise FileNotFoundError(f"Папка склада не найдена: {sklad_dir}")
    files = [
        p
        for p in sklad_dir.iterdir()
        if p.is_file()
        and p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
        and is_filled_sklad_template_name(p.name)
    ]
    if not files:
        raise FileNotFoundError(
            f"В {sklad_dir} нет заполненного «Шаблон ДД.ММ.ГГ». "
            "Сначала «Шаблон → копия на дату», затем заполните сетку."
        )
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0].resolve()


def parse_sklad_template(path: Path) -> TemplateDemand:
    """
    Первая таблица «Эквадор»: строки до «Итог».
    Колонки длин — из строки заголовка (F=50, G=60, …).
    Код сорта: колонка B или C (в новых шаблонах «Эквадор» в B, в старых — в C).
    """
    path = path.resolve()
    suffix = path.suffix.lower()
    if suffix == ".xls":
        raise ValueError("Шаблон .xls не поддерживается — сохраните как .xlsx")

    rows = grid_by_row(read_excel_grid(path))
    header_row: int | None = None
    code_col: str = "C"
    length_cols: dict[str, str] = {}  # length label → col letter

    for row_no in sorted(rows):
        cells = rows[row_no]
        ecuador_col: str | None = None
        for col in ("B", "C"):
            if _norm(cells.get(col, "")).casefold() == "эквадор":
                ecuador_col = col
                break
        if ecuador_col is None:
            continue
        found: dict[str, str] = {}
        for col, val in cells.items():
            label = _norm(val)
            if label in _TARGET_LENGTHS or label in {"70", "80", "40", "90", "100"}:
                found[label] = col
        if "50" in found or "60" in found:
            header_row = row_no
            code_col = ecuador_col
            length_cols = found
            break

    if header_row is None:
        raise RuntimeError(f"В {path.name} нет таблицы Эквадор с длинами 50/60")

    lines: list[TemplateLine] = []
    totals: dict[str, int] = {lab: 0 for lab in _TARGET_LENGTHS}
    for row_no in range(header_row + 1, max(rows) + 1):
        cells = rows.get(row_no, {})
        d = _norm(cells.get("D", ""))
        c = _norm(cells.get(code_col, ""))
        if not c and code_col == "B":
            c = _norm(cells.get("C", ""))
        elif not c and code_col == "C":
            c = _norm(cells.get("B", ""))
        if d.casefold() == "итог":
            for lab in _TARGET_LENGTHS:
                col = length_cols.get(lab)
                if col:
                    totals[lab] = _as_qty(cells.get(col, ""))
            break
        if not c and not d:
            continue
        if not c:
            continue
        qtys = {
            lab: _as_qty(cells.get(length_cols[lab], ""))
            for lab in _TARGET_LENGTHS
            if lab in length_cols
        }
        lines.append(
            TemplateLine(
                excel_row=row_no,
                code=c,
                title_ru=d,
                qtys=qtys,
            )
        )

    return TemplateDemand(path=path, lines=tuple(lines), totals=totals)


def _biflorica_header(rows: dict[int, dict[str, str]], path: Path) -> tuple[int, dict[str, str]]:
    return biflorica_deals_header_or_raise(rows, path=path)


def _is_mix_rose(typ: str, variety: str) -> bool:
    t = _norm(typ).casefold()
    v = _norm(variety).casefold()
    if "роз" not in t:
        return False
    # Сорт Mix / Микс (не MixAlstromeria и т.п.)
    return v in {"mix", "микс"}


def _priced_length_labels(
    cells: dict[str, str],
    length_map: dict[str, str],
) -> list[str]:
    labels: list[tuple[int, str]] = []
    for lab, col in length_map.items():
        price = _as_float(cells.get(col, ""))
        if price is not None and price > 0:
            labels.append((_col_to_index(col), lab))
    labels.sort()
    return [lab for _, lab in labels]


def _stems_for_length(
    cells: dict[str, str],
    length: str,
    length_map: dict[str, str],
) -> int:
    """Стебли строки для одной длины (учёт «75|25» при нескольких ценах)."""
    priced = _priced_length_labels(cells, length_map)
    if length not in priced:
        return 0
    o_raw = _norm(cells.get("O", ""))
    if not o_raw:
        return 0
    if len(priced) == 1:
        return _as_qty(o_raw)
    parts = [p.strip() for p in o_raw.split("|")]
    if len(parts) == len(priced):
        return _as_qty(parts[priced.index(length)])
    if priced[0] == length:
        return _as_qty(o_raw)
    return 0


@dataclass(frozen=True)
class MixPoolRow:
    row_no: int
    plantation: str
    price: float
    stems: int


def scan_mix_pool(
    rows: dict[int, dict[str, str]],
    header_row: int,
    length_map: dict[str, str],
    length: str,
) -> list[MixPoolRow]:
    price_col = length_map.get(length)
    if not price_col:
        return []
    pool: list[MixPoolRow] = []
    for row_no in sorted(rows):
        if row_no <= header_row:
            continue
        cells = rows[row_no]
        if not _is_mix_rose(cells.get("C", ""), cells.get("D", "")):
            continue
        price = _as_float(cells.get(price_col, ""))
        if price is None or price <= 0:
            continue
        stems = _stems_for_length(cells, length, length_map)
        if stems <= 0:
            continue
        pool.append(
            MixPoolRow(
                row_no=row_no,
                plantation=_norm(cells.get("B", "")),
                price=price,
                stems=stems,
            )
        )
    return pool


def log_mix_capacity(
    demand: TemplateDemand,
    biflorica_path: Path,
    rows: dict[int, dict[str, str]],
    header_row: int,
    length_map: dict[str, str],
    *,
    log: LogFn | None = None,
) -> None:
    _lg = log or _default_log
    try:
        st = biflorica_path.stat()
        mtime = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
        _lg(
            f"Миксы: Biflorica {biflorica_path.name} "
            f"({st.st_size} байт, изменён {mtime})"
        )
    except OSError:
        _lg(f"Миксы: Biflorica {biflorica_path.name}")
    _lg(f"Миксы: шаблон {demand.path.name}, нужно {demand.totals}")
    for length in _TARGET_LENGTHS:
        need = demand.totals.get(length, 0)
        if need <= 0:
            continue
        pool = scan_mix_pool(rows, header_row, length_map, length)
        available = sum(r.stems for r in pool)
        _lg(f"Миксы: в файле Mix {length} см — {available} шт ({len(pool)} строк)")
        for row in pool:
            plant = row.plantation or "-"
            _lg(
                f"Миксы:   строка {row.row_no} {plant}: "
                f"{row.stems} шт × {row.price:.4f}"
            )
        if available < need:
            _lg(
                f"Миксы: ⚠ не хватает {need - available} шт на {length} см "
                f"(нужно {need}, есть {available})"
            )


def plan_mix_allocation(
    demand: TemplateDemand,
    biflorica_path: Path,
    *,
    log: LogFn | None = None,
) -> list[LengthPlan]:
    _lg = log or _default_log
    rows = grid_by_row(read_excel_grid(biflorica_path))
    header_row, length_map = _biflorica_header(rows, biflorica_path)
    plans: list[LengthPlan] = []

    for length in _TARGET_LENGTHS:
        need = demand.totals.get(length, 0)
        filled = [ln for ln in demand.lines if ln.qtys.get(length, 0) > 0]
        if need <= 0 or not filled:
            _lg(f"Миксы: длина {length} — пропуск (need={need}, строк={len(filled)})")
            continue
        sum_lines = sum(ln.qtys.get(length, 0) for ln in filled)
        if sum_lines != need:
            _lg(
                f"Миксы: длина {length} — итог {need}, сумма строк {sum_lines} "
                "(беру итог из шаблона)"
            )

        price_col = length_map.get(length)
        if not price_col:
            raise RuntimeError(f"В Biflorica нет колонки длины {length}")

        candidates: list[tuple[float, int, int, str, float]] = []
        pool = scan_mix_pool(rows, header_row, length_map, length)
        for row in pool:
            candidates.append(
                (row.price, row.stems, row.row_no, row.plantation, row.price)
            )
        candidates.sort(key=lambda x: (-x[0], x[2]))  # дорогие сначала

        available = sum(avail for _, avail, _, _, _ in candidates)
        left = need
        takes: list[MixTake] = []
        for price, avail, row_no, plant, _ in candidates:
            if left <= 0:
                break
            take = min(avail, left)
            takes.append(
                MixTake(
                    bif_row=row_no,
                    take_qty=take,
                    price=price,
                    plantation=plant,
                )
            )
            left -= take

        if left > 0:
            rows_hint = "; ".join(
                f"#{r.row_no} {r.plantation or '-'}:{r.stems}"
                for r in pool[:8]
            )
            if len(pool) > 8:
                rows_hint += f" … (+{len(pool) - 8})"
            hint = ""
            if available > 0 and available < need:
                hint = (
                    " Если миксы уже запускали — восстановите файл из «… до миксов ….xlsx» "
                    "или скачайте Biflorica заново и убедитесь, что в папке один свежий файл."
                )
            raise RuntimeError(
                f"Миксы: для длины {length} нужно {need}, в Biflorica Mix есть "
                f"только {available} (не хватает {left}). Строки Mix: {rows_hint}.{hint}"
            )

        total_stems = sum(t.take_qty for t in takes)
        avg = sum(t.take_qty * t.price for t in takes) / total_stems
        plan = LengthPlan(
            length=length,
            need=need,
            lines=filled,
            takes=takes,
            avg_price=avg,
        )
        plans.append(plan)
        _lg(
            f"Миксы: {length} см — need={need}, позиций={len(filled)}, "
            f"источников Mix={len(takes)}, avg={avg:.4f}"
        )
        for t in takes:
            _lg(
                f"Миксы: {length} см ← строка {t.bif_row} "
                f"{t.plantation or '(без плантации)'}: {t.take_qty} шт × {t.price:.4f}"
            )
        _lg(
            f"Миксы: {length} см — проверка: сумма взятого {total_stems} шт, "
            f"стоимость {sum(t.take_qty * t.price for t in takes):.2f}"
        )

    return plans


def backup_biflorica_before_mixes(path: Path) -> Path:
    """Копия «<имя> до миксов ГГГГ-ММ-ДД ЧЧММСС.xlsx» рядом с файлом."""
    stamp = datetime.now().strftime("%Y-%m-%d %H%M%S")
    backup = path.with_name(f"{path.stem} до миксов {stamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


_BIFLORICA_MONEY_FORMAT = '#,##0.00_- "$"'
_PLANTATION_EMPTY = "-"


def _sample_money_format(
    ws: object,
    col_index: dict[str, int],
    length_map: dict[str, str],
    *,
    header_row: int,
) -> str:
    """Формат цены/суммы из существующих строк отчёта."""
    letters = list(length_map.values())
    if "P" in col_index:
        letters.append("P")
    for row_no in range(header_row + 1, ws.max_row + 1):  # type: ignore[attr-defined]
        for letter in letters:
            col = col_index.get(letter)
            if col is None:
                continue
            cell = ws.cell(row_no, col)  # type: ignore[attr-defined]
            if cell.value is not None and cell.number_format and cell.number_format != "General":
                return cell.number_format
    return _BIFLORICA_MONEY_FORMAT


def _set_money_cell(cell: object, value: float, number_format: str) -> None:
    cell.value = round(value, 2)  # type: ignore[attr-defined]
    cell.number_format = number_format  # type: ignore[attr-defined]


def apply_mix_plans_to_biflorica(
    biflorica_path: Path,
    plans: list[LengthPlan],
    *,
    log: LogFn | None = None,
) -> Path:
    """
    Вносит изменения в Biflorica.xlsx:
    - уменьшает/удаляет строки Mix, с которых брали;
    - добавляет новые строки без плантации, сорт = код из шаблона (C),
      цена = avg, кол-во = из шаблона.
    """
    _lg = log or _default_log
    path = ensure_xlsx_workbook(biflorica_path.resolve())
    backup = backup_biflorica_before_mixes(path)
    _lg(f"Миксы: резервная копия → {backup.name}")
    wb = load_workbook(path)
    ws = pick_data_worksheet(wb)

    # карта: длина → индекс колонки 1-based
    rows_grid = grid_by_row(read_excel_grid(path))
    _header_row, length_map = _biflorica_header(rows_grid, path)
    col_index = {letter: _col_to_index(letter) for letter in set(length_map.values()) | {"B", "C", "D", "O", "P"}}
    money_fmt = _sample_money_format(ws, col_index, length_map, header_row=_header_row)

    # 1) списать qty с Mix-строк (накопить take по строке)
    take_by_row: dict[int, int] = {}
    for plan in plans:
        for t in plan.takes:
            take_by_row[t.bif_row] = take_by_row.get(t.bif_row, 0) + t.take_qty

    rows_to_delete: list[int] = []
    for row_no, take in take_by_row.items():
        o_col = col_index["O"]
        cur = _as_qty(ws.cell(row_no, o_col).value)
        new_qty = cur - take
        if new_qty <= 0:
            rows_to_delete.append(row_no)
        else:
            ws.cell(row_no, o_col).value = new_qty
            # пересчёт суммы сделки если есть цена в одной длине
            price = None
            for lab, letter in length_map.items():
                p = _as_float(ws.cell(row_no, col_index[letter]).value)
                if p is not None and p > 0:
                    price = p
                    break
            if price is not None and "P" in col_index:
                _set_money_cell(ws.cell(row_no, col_index["P"]), price * new_qty, money_fmt)

    for row_no in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_no, 1)
        _lg(f"Миксы: удалена строка Mix #{row_no}")

    # после delete индексы сдвинулись — новые строки только в конец
    # 2) добавить позиции
    for plan in plans:
        price_letter = length_map[plan.length]
        price_col = col_index[price_letter]
        for line in plan.lines:
            qty = line.qtys.get(plan.length, 0)
            if qty <= 0:
                continue
            new_row = ws.max_row + 1
            ws.cell(new_row, col_index["B"]).value = _PLANTATION_EMPTY
            ws.cell(new_row, col_index["C"]).value = "Роза"
            ws.cell(new_row, col_index["D"]).value = line.code
            for letter in length_map.values():
                ws.cell(new_row, col_index[letter]).value = None
            _set_money_cell(ws.cell(new_row, price_col), plan.avg_price, money_fmt)
            ws.cell(new_row, col_index["O"]).value = qty
            if "P" in col_index:
                _set_money_cell(
                    ws.cell(new_row, col_index["P"]),
                    plan.avg_price * qty,
                    money_fmt,
                )
            _lg(
                f"Миксы: + {line.code} / {plan.length} см × {qty} @ {plan.avg_price:.4f}"
            )

    wb.save(path)
    wb.close()
    _lg(f"Миксы: сохранено → {path.name}")
    return path


def _col_to_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def run_mix_separation(
    *,
    template_path: Path,
    biflorica_path: Path,
    log: LogFn | None = None,
) -> tuple[TemplateDemand, list[LengthPlan], Path]:
    _lg = log or _default_log
    demand = parse_sklad_template(template_path)
    _lg(
        f"Миксы: шаблон {demand.path.name}, строк={len(demand.lines)}, "
        f"итоги={demand.totals}"
    )
    rows = grid_by_row(read_excel_grid(biflorica_path))
    header_row, length_map = _biflorica_header(rows, biflorica_path)
    log_mix_capacity(
        demand,
        biflorica_path,
        rows,
        header_row,
        length_map,
        log=_lg,
    )
    plans = plan_mix_allocation(demand, biflorica_path, log=_lg)
    if not plans:
        raise RuntimeError("Миксы: в шаблоне нет заполненных 50/60 для разбора")
    out = apply_mix_plans_to_biflorica(biflorica_path, plans, log=_lg)
    return demand, plans, out


def run_mix_separation_from_dirs(
    sklad_dir: Path,
    biflorica_dir: Path,
    *,
    template_path: Path | None = None,
    biflorica_path: Path | None = None,
    log: LogFn | None = None,
) -> tuple[Path, Path, list[LengthPlan]]:
    _lg = log or _default_log
    tpl = template_path or find_latest_filled_sklad_template(sklad_dir)
    bif = ensure_xlsx_workbook(
        (biflorica_path or find_latest_biflorica_report(biflorica_dir)).resolve()
    )
    if is_split_output_name(bif.name):
        raise ValueError(f"Нельзя разбирать уже разделённый файл: {bif.name}")
    _lg(f"Миксы: шаблон={tpl.name}, biflorica={bif.name}")
    _demand, plans, out = run_mix_separation(
        template_path=tpl,
        biflorica_path=bif,
        log=_lg,
    )
    return tpl, out, plans
