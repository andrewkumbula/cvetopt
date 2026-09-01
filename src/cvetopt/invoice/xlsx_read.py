from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import xlrd

_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_COL_RE = re.compile(r"^([A-Z]+)(\d+)$")
_OLE_MAGIC = b"\xd0\xcf\x11\xe0"
_BIFLORICA_MARKER = "плантация"


def _col_index_to_letter(idx: int) -> str:
    n = idx + 1
    letters = ""
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _col_letter_to_index(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _cell_to_text(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return " ".join(str(value).split()).strip()


def excel_file_kind(path: Path) -> str:
    """xlsx | xls | invalid"""
    try:
        head = path.read_bytes()[:4]
    except OSError:
        return "invalid"
    if head[:2] == b"PK":
        return "xlsx"
    if head == _OLE_MAGIC:
        return "xls"
    return "invalid"


def _xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    shared: list[str] = []
    for si in root.findall(".//m:si", _NS):
        shared.append("".join((n.text or "") for n in si.iter()))
    return shared


def _xlsx_cell_text(cell: ET.Element, shared: list[str]) -> str:
    cell_type = cell.get("t")
    if cell_type == "inlineStr":
        is_node = cell.find("m:is", _NS)
        if is_node is not None:
            return _cell_to_text("".join((n.text or "") for n in is_node.iter()))
        return ""
    v = cell.find("m:v", _NS)
    if v is None or v.text is None:
        is_node = cell.find("m:is", _NS)
        if is_node is not None:
            return _cell_to_text("".join((n.text or "") for n in is_node.iter()))
        return ""
    if cell_type == "s":
        try:
            return _cell_to_text(shared[int(v.text)])
        except (IndexError, ValueError):
            return _cell_to_text(v.text)
    if cell_type == "b":
        return "TRUE" if v.text == "1" else "FALSE"
    return _cell_to_text(v.text)


def _parse_xlsx_sheet_xml(root: ET.Element, shared: list[str]) -> dict[str, str]:
    """Парсит лист; поддерживает ячейки без атрибута r (как в отчётах Biflorica)."""
    grid: dict[str, str] = {}
    for row_el in root.findall(".//m:sheetData/m:row", _NS):
        row_num = int(row_el.get("r", "0") or "0")
        col_idx = 0
        for cell in row_el.findall("m:c", _NS):
            ref = cell.get("r")
            if ref:
                m = _COL_RE.match(ref)
                if not m:
                    continue
                col_letter, row_num = m.group(1), int(m.group(2))
                col_idx = _col_letter_to_index(col_letter) + 1
            else:
                if row_num <= 0:
                    continue
                col_letter = _col_index_to_letter(col_idx)
                ref = f"{col_letter}{row_num}"
                col_idx += 1
            text = _xlsx_cell_text(cell, shared)
            if text:
                grid[ref] = text
    return grid


def _grid_has_biflorica_marker(grid: dict[str, str]) -> bool:
    for val in grid.values():
        if _BIFLORICA_MARKER in _cell_to_text(val).casefold():
            return True
    return False


def _pick_best_grid(candidates: list[dict[str, str]]) -> dict[str, str]:
    if not candidates:
        return {}
    marked = [g for g in candidates if _grid_has_biflorica_marker(g)]
    pool = marked or candidates
    return max(pool, key=len)


def read_xls_grid(path: Path) -> dict[str, str]:
    """Все листы .xls → лист с данными (приоритет — «ПЛАНТАЦИЯ»)."""
    book = xlrd.open_workbook(str(path))
    candidates: list[dict[str, str]] = []
    for sheet in book.sheets():
        grid: dict[str, str] = {}
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                text = _cell_to_text(sheet.cell_value(row_idx, col_idx))
                if not text:
                    continue
                grid[f"{_col_index_to_letter(col_idx)}{row_idx + 1}"] = text
        if grid:
            candidates.append(grid)
    return _pick_best_grid(candidates)


def read_xlsx_grid_xml(path: Path) -> dict[str, str]:
    """Быстрый разбор xlsx через XML (все листы)."""
    with zipfile.ZipFile(path) as zf:
        shared = _xlsx_shared_strings(zf)
        sheet_names = sorted(
            n
            for n in zf.namelist()
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
        )
        if not sheet_names:
            return {}
        candidates: list[dict[str, str]] = []
        for sheet_name in sheet_names:
            root = ET.fromstring(zf.read(sheet_name))
            grid = _parse_xlsx_sheet_xml(root, shared)
            if grid:
                candidates.append(grid)
        return _pick_best_grid(candidates)


def read_xlsx_grid_openpyxl(path: Path) -> dict[str, str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    candidates: list[dict[str, str]] = []
    try:
        for ws in wb.worksheets:
            grid: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    text = _cell_to_text(cell.value)
                    if text and cell.coordinate:
                        grid[cell.coordinate] = text
            if grid:
                candidates.append(grid)
    finally:
        wb.close()
    return _pick_best_grid(candidates)


def read_xlsx_grid(path: Path) -> dict[str, str]:
    """xlsx → «A1» → значение; XML, при пустом результате — openpyxl."""
    grid = read_xlsx_grid_xml(path)
    if grid:
        return grid
    return read_xlsx_grid_openpyxl(path)


def read_excel_grid(path: Path) -> dict[str, str]:
    """Читает .xlsx или старый .xls (в т.ч. .xls под именем .xlsx)."""
    kind = excel_file_kind(path)
    if kind == "xlsx":
        return read_xlsx_grid(path)
    if kind == "xls":
        return read_xls_grid(path)
    raise ValueError(f"Не Excel или повреждённый файл: {path.name}")


def excel_grid_stats(path: Path) -> str:
    """Краткая сводка для диагностики (размер, формат, число ячеек)."""
    try:
        size = path.stat().st_size
    except OSError as exc:
        return str(exc)
    kind = excel_file_kind(path)
    try:
        grid = read_excel_grid(path)
        cells = len(grid)
        marker = "да" if _grid_has_biflorica_marker(grid) else "нет"
        return f"{size} байт, формат {kind}, ячеек {cells}, ПЛАНТАЦИЯ: {marker}"
    except Exception as exc:
        return f"{size} байт, формат {kind}, ошибка чтения: {exc}"


def ensure_xlsx_workbook(path: Path) -> Path:
    """
    Если файл — старый .xls (даже с расширением .xlsx), пересохраняет как настоящий .xlsx.
    Нужно для openpyxl (миксы, сплит).
    """
    from openpyxl import Workbook

    path = path.resolve()
    if excel_file_kind(path) != "xls":
        return path
    grid = read_xls_grid(path)
    if not grid:
        raise ValueError(f"В {path.name} нет данных для конвертации в xlsx")
    wb = Workbook()
    ws = wb.active
    for ref, val in grid.items():
        if _COL_RE.match(ref):
            ws[ref] = val
    wb.save(path)
    wb.close()
    return path


def grid_by_row(grid: dict[str, str]) -> dict[int, dict[str, str]]:
    rows: dict[int, dict[str, str]] = {}
    for ref, val in grid.items():
        m = _COL_RE.match(ref)
        if not m:
            continue
        col, row = m.group(1), int(m.group(2))
        rows.setdefault(row, {})[col] = val
    return rows
