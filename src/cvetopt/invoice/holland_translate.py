from __future__ import annotations

from collections.abc import Callable
from datetime import date
from pathlib import Path

from cvetopt.invoice.description_dictionary import (
    load_description_dictionary,
    lookup_translation,
)

LogFn = Callable[[str], None]


def _default_log(_msg: str) -> None:
    pass


def _next_col_letter(col: str) -> str:
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - 64)
    n += 1
    out: list[str] = []
    while n:
        n, rem = divmod(n - 1, 26)
        out.append(chr(65 + rem))
    return "".join(reversed(out))


def holland_export_candidates(output_dir: Path, on_date: date | None = None) -> list[Path]:
    if not output_dir.is_dir():
        return []
    names: list[str] = ["Голландия_1_*.xlsx"]
    if on_date is not None:
        names.insert(0, f"Голландия_1_{on_date.strftime('%d.%m.%Y')}.xlsx")
        names.insert(1, f"Голландия_1_{on_date.day}.{on_date.month}.{on_date.year}.xlsx")
    seen: set[Path] = set()
    found: list[Path] = []
    for pattern in names:
        for p in output_dir.glob(pattern):
            rp = p.resolve()
            if rp not in seen and rp.is_file():
                seen.add(rp)
                found.append(rp)
    if found:
        return sorted(found, key=lambda p: p.stat().st_mtime, reverse=True)
    return sorted(output_dir.glob("Голландия_1_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)


def find_holland_export_file(
    output_dir: Path,
    *,
    on_date: date | None = None,
) -> Path | None:
    items = holland_export_candidates(output_dir, on_date=on_date)
    return items[0] if items else None


def _find_description_columns(header: dict[str, str]) -> tuple[str, str] | None:
    for col, val in header.items():
        if str(val).strip().casefold() == "description":
            return col, _next_col_letter(col)
    return None


def translate_holland_export(
    export_path: Path,
    dictionary_path: Path,
    *,
    log: LogFn | None = None,
) -> tuple[int, int, int]:
    """
    Заполняет столбец справа от Description переводом из словаря.
    Возвращает (переведено, без_перевода, всего_строк_данных).
    """
    import openpyxl

    _lg = log or _default_log
    export_path = export_path.resolve()
    dictionary = load_description_dictionary(dictionary_path)
    if not dictionary:
        raise RuntimeError(f"Словарь пуст: {dictionary_path}")

    _lg(f"Словарь: {len(dictionary)} записей из {dictionary_path.name}")

    wb = openpyxl.load_workbook(export_path)
    ws = wb.active

    header: dict[str, str] = {}
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip():
            header[cell.column_letter] = str(cell.value).strip()

    cols = _find_description_columns(header)
    if cols is None:
        wb.close()
        raise RuntimeError(
            f"В {export_path.name} нет заголовка Description в первой строке."
        )
    desc_col, trans_col = cols
    _lg(f"Description → колонка {trans_col} (рядом с {desc_col})")

    translated = 0
    missing = 0
    total = 0
    for row in range(2, ws.max_row + 1):
        desc_cell = ws[f"{desc_col}{row}"]
        raw = desc_cell.value
        if raw is None or str(raw).strip() == "":
            continue
        total += 1
        text = str(raw).strip()
        tr = lookup_translation(dictionary, text)
        target = ws[f"{trans_col}{row}"]
        if tr:
            target.value = tr
            translated += 1
        else:
            target.value = None
            missing += 1

    wb.save(export_path)
    wb.close()
    _lg(
        f"Перевод в {export_path.name}: строк {total}, "
        f"переведено {translated}, без перевода {missing}."
    )
    return translated, missing, total


def postprocess_holland_after_auto1(
    *,
    sklad_output_dir: Path,
    dictionary_path: Path,
    on_date: date | None = None,
    log: LogFn | None = None,
) -> Path | None:
    """Ищет свежий Голландия_1_*.xlsx в папке склада и переводит Description."""
    _lg = log or _default_log
    export_file = find_holland_export_file(sklad_output_dir, on_date=on_date)
    if export_file is None:
        _lg(f"Файл Голландия_1_*.xlsx не найден в {sklad_output_dir}")
        return None
    translate_holland_export(export_file, dictionary_path, log=_lg)
    return export_file
