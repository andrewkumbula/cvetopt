"""
Разделение отчёта Biflorica по типу цветка (аналог макроса «Гипсофила» + «Да»).

Создаёт рядом с исходником:
  <имя> Гипсофила.xlsx
  <имя> Роза.xlsx
Исходный полный файл не удаляет.
"""

from __future__ import annotations

import re
import shutil
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from cvetopt.core.runtime_settings import BIFLORICA_DOWNLOAD_PREFIX, order_id_from_biflorica_report
from cvetopt.invoice.xlsx_read import grid_by_row, read_xlsx_grid

LogFn = Callable[[str], None]

_SPLIT_SUFFIXES = ("Гипсофила", "Роза", "Прочее", "Гортензия")
_TYPE_COL = "C"
_DATA_FIRST_FALLBACK = 7
_BIFLORICA_LENGTHS = frozenset({"40", "50", "60", "70", "80", "90", "100", "100+"})


def _default_log(_msg: str) -> None:
    pass


def _norm(text: object) -> str:
    return " ".join(str(text or "").split()).strip()


def is_split_output_name(name: str) -> bool:
    """Уже разделённый файл: «… Гипсофила.xlsx» / «… Роза.xlsx»."""
    stem = Path(name).stem
    for suf in _SPLIT_SUFFIXES:
        if stem.endswith(f" {suf}") or stem.casefold().endswith(f" {suf.casefold()}"):
            return True
    return False


def classify_flower_type(type_name: str) -> str | None:
    """
    Возвращает ярлык файла-выгрузки или None (остаётся только в полном отчёте).
    Гипсофила → «Гипсофила»; любая роза → «Роза».
    """
    t = _norm(type_name).casefold()
    if not t:
        return None
    if "гипсофил" in t:
        return "Гипсофила"
    if "роз" in t:  # Роза, Крашеная роза
        return "Роза"
    return None


def normalize_biflorica_length_label(value: object) -> str | None:
    text = _norm(value).replace(",", ".")
    if not text:
        return None
    if text in _BIFLORICA_LENGTHS:
        return text
    try:
        num = float(text)
        if num == int(num):
            label = str(int(num))
            if label in _BIFLORICA_LENGTHS:
                return label
    except ValueError:
        pass
    return None


def _biflorica_header_marker_row(row: dict[str, str]) -> bool:
    a = _norm(row.get("A", "")).casefold()
    b = _norm(row.get("B", "")).casefold()
    return a == "дата и время сделки" or b == "плантация"


def find_biflorica_deals_header(
    rows: dict[int, dict[str, str]],
) -> tuple[int, dict[str, str]] | None:
    """Строка «ПЛАНТАЦИЯ» + колонки длин 40…100+ или None."""
    for row_no in sorted(rows):
        row = rows[row_no]
        if not _biflorica_header_marker_row(row):
            continue
        length_map: dict[str, str] = {}
        for col, val in row.items():
            lab = normalize_biflorica_length_label(val)
            if lab:
                length_map[lab] = col
        if length_map:
            return row_no, length_map
    return None


def diagnose_biflorica_report(path: Path) -> str | None:
    """None — файл пригоден; иначе короткая причина на русском."""
    if not path.is_file():
        return "файл не найден"
    try:
        size = path.stat().st_size
    except OSError as exc:
        return str(exc)
    if size < 512:
        return f"слишком маленький ({size} байт) — вероятно пустой или обрыв скачивания"
    try:
        head = path.read_bytes()[:4]
    except OSError as exc:
        return str(exc)
    if head[:2] != b"PK":
        return (
            "не настоящий xlsx (не ZIP) — откройте в Excel: если не открывается, "
            "удалите и скачайте отчёт заново кнопкой Biflorica"
        )
    try:
        rows = grid_by_row(read_xlsx_grid(path))
    except zipfile.BadZipFile:
        return "повреждённый xlsx — удалите и скачайте отчёт заново"
    except (OSError, ValueError, KeyError) as exc:
        return f"не удалось прочитать: {exc}"
    if not rows:
        return "пустой лист Excel"
    if find_biflorica_deals_header(rows) is not None:
        return None

    marker_rows: list[tuple[int, dict[str, str]]] = []
    for row_no in sorted(rows):
        row = rows[row_no]
        if _biflorica_header_marker_row(row):
            marker_rows.append(
                (row_no, {col: _norm(val) for col, val in row.items() if _norm(val)})
            )
    if not marker_rows:
        preview: list[str] = []
        for row_no in sorted(rows)[:6]:
            a = _norm(rows[row_no].get("A", ""))
            if a:
                preview.append(a[:80])
        tail = "; ".join(preview) if preview else "(пусто)"
        return (
            "нет таблицы сделок (строка «ПЛАНТАЦИЯ» и колонки 40–100). "
            f"Начало файла: {tail}"
        )

    row_no, cols = marker_rows[0]
    col_preview = ", ".join(f"{c}={v!r}" for c, v in sorted(cols.items())[:10])
    return (
        f"строка заголовка есть (строка {row_no}), но нет колонок длин 40–100. "
        f"Заголовок: {col_preview}"
    )


def biflorica_deals_header_or_raise(
    rows: dict[int, dict[str, str]],
    *,
    path: Path | None = None,
) -> tuple[int, dict[str, str]]:
    found = find_biflorica_deals_header(rows)
    if found is not None:
        return found
    name = path.name if path is not None else "Biflorica"
    raise RuntimeError(
        f"В {name} нет таблицы сделок Biflorica (строка «ПЛАНТАЦИЯ» и колонки 40–100). "
        "Нужен полный отчёт BiFlorica-*.xlsx после скачивания, "
        "не «… Роза.xlsx» / «… Гипсофила.xlsx»."
    )


def is_biflorica_deals_report(path: Path) -> bool:
    return diagnose_biflorica_report(path) is None


def _find_header_row(rows: dict[int, dict[str, str]]) -> int:
    found = find_biflorica_deals_header(rows)
    if found is not None:
        return found[0]
    for row_no in sorted(rows):
        row = rows[row_no]
        if _biflorica_header_marker_row(row):
            return row_no
    return _DATA_FIRST_FALLBACK - 1


def find_latest_biflorica_report(download_dir: Path) -> Path:
    """Самый свежий полный BiFlorica-*.xlsx в корне папки (не архив, не сплит)."""
    if not download_dir.is_dir():
        raise FileNotFoundError(f"Папка Biflorica не найдена: {download_dir}")

    files: list[Path] = []
    for path in download_dir.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() != ".xlsx":
            continue
        if is_split_output_name(path.name):
            continue
        name = path.name
        if not (
            name.startswith(BIFLORICA_DOWNLOAD_PREFIX)
            or name.lower().startswith("biflorica")
            or re.match(r"^\d+__", path.stem)
        ):
            continue
        files.append(path)

    if not files:
        raise FileNotFoundError(
            f"В {download_dir} нет полного отчёта BiFlorica-*.xlsx. "
            "Сначала скачайте отчёт или укажите файл явно."
        )
    files.sort(
        key=lambda p: (
            0 if order_id_from_biflorica_report(p) else 1,
            -p.stat().st_mtime,
        )
    )
    checked: list[str] = []
    for path in files:
        reason = diagnose_biflorica_report(path)
        if reason is None:
            return path.resolve()
        checked.append(f"{path.name} ({reason})")
    preview = "; ".join(checked[:3])
    if len(checked) > 3:
        preview += f" … (+{len(checked) - 3})"
    raise FileNotFoundError(
        f"В {download_dir} нет пригодного отчёта Biflorica. "
        f"Проверены: {preview}. "
        "Нужен полный BiFlorica-*.xlsx со строкой «ПЛАНТАЦИЯ» и колонками 40–100 — "
        "скачайте заново кнопкой Biflorica (не «… Роза.xlsx» / «… Гипсофила.xlsx»)."
    )


@dataclass(frozen=True)
class BifloricaSplitResult:
    source: Path
    outputs: dict[str, Path]  # label → path
    counts: dict[str, int]  # label → data rows


def split_biflorica_by_type(
    source: Path,
    *,
    output_dir: Path | None = None,
    log: LogFn | None = None,
) -> BifloricaSplitResult:
    """
    Копирует source в «… Гипсофила.xlsx» / «… Роза.xlsx», оставляя в каждом
    шапку + строки нужного типа (колонка C «ТИП»).
    """
    _lg = log or _default_log
    source = source.resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Файл не найден: {source}")
    if is_split_output_name(source.name):
        raise ValueError(f"Файл уже выглядит как результат сплита: {source.name}")

    out_dir = (output_dir or source.parent).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    grid = read_xlsx_grid(source)
    rows = grid_by_row(grid)
    header_row = _find_header_row(rows)
    _lg(f"Гипсофила: источник {source.name}, шапка строка {header_row}")

    # 1-based Excel rows belonging to each label
    by_label: dict[str, list[int]] = {"Гипсофила": [], "Роза": []}
    other = 0
    for row_no in sorted(rows):
        if row_no <= header_row:
            continue
        typ = rows[row_no].get(_TYPE_COL, "")
        if not _norm(typ) and not _norm(rows[row_no].get("B", "")):
            continue
        label = classify_flower_type(typ)
        if label is None:
            other += 1
            continue
        by_label[label].append(row_no)

    outputs: dict[str, Path] = {}
    counts: dict[str, int] = {}

    for label, keep_rows in by_label.items():
        dest = out_dir / f"{source.stem} {label}.xlsx"
        if dest.exists():
            dest.unlink()
        shutil.copy2(source, dest)
        _filter_workbook_rows(dest, header_row=header_row, keep_data_rows=set(keep_rows))
        outputs[label] = dest
        counts[label] = len(keep_rows)
        _lg(f"Гипсофила: {label} → {dest.name} ({len(keep_rows)} строк)")

    if other:
        _lg(f"Гипсофила: прочих типов (только в полном файле): {other}")

    return BifloricaSplitResult(source=source, outputs=outputs, counts=counts)


def _filter_workbook_rows(
    path: Path,
    *,
    header_row: int,
    keep_data_rows: set[int],
) -> None:
    """Удаляет строки данных, не входящие в keep_data_rows (1-based)."""
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    # Удаляем снизу вверх, чтобы индексы не съезжали.
    max_row = ws.max_row or header_row
    for row_no in range(max_row, header_row, -1):
        if row_no not in keep_data_rows:
            ws.delete_rows(row_no, 1)
    wb.save(path)
    wb.close()


def run_gypsophila_split(
    download_dir: Path,
    *,
    source: Path | None = None,
    log: LogFn | None = None,
) -> BifloricaSplitResult:
    """Находит свежий Biflorica (или берёт source) и пишет Гипсофила/Роза."""
    _lg = log or _default_log
    src = source.resolve() if source is not None else find_latest_biflorica_report(download_dir)
    _lg(f"Гипсофила: обрабатываю {src}")
    return split_biflorica_by_type(src, output_dir=src.parent, log=_lg)
